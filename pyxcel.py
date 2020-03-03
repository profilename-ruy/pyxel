import os
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Definitions
wb = Workbook()
ws = wb.active
list_names = glob.glob("*.txt")
list_length = len(list_names)
file_name= input("Defina o nome da planilha:")

# Get sheet name
def get_file(x):
	name = list_names[x]
	name = name.strip(".txt")
	return name
	
# Open File
def open_file(y):
	file = open(list_names[y],"r")
	file = file.readlines()	
	return file

# Create a New Sheet
def sheet_name(z):
	name = get_file(z)
	wb.create_sheet(name)	

# Save archive
def save():
	wb.save(file_name + ".xlsx")

# Get IP List
def get_IP(j):
	IP = []
	for x in j:
		x = x.split("\t")
		IP.append(x[0])
	return IP

# Get date
def get_date(i):
	date = []
	for x in i:
		x = x.split('\t')
		del x[0]
		for g in x:
			g = g.split('"')
			date.append(g[0])
	return date	

# Get Status
def get_status(c):
	status = []
	for x in c:
		x = x.split('\t')
		del x[0]
		for j in x:
			j = j.split('"')
			status.append(j[1])
	return status	

# Active a Sheet
def act_sheet(h):
	worksheet_names = wb.sheetnames 
	sheet_index = worksheet_names.index(h)
	wb.active = sheet_index

# Set column dimensions
def column_dimensions():
	wb.active.column_dimensions['A'].width = 13.4
	wb.active.column_dimensions['B'].width = 25.7
	wb.active.column_dimensions['C'].width = 49.9

def title():
	wb.active.insert_rows(1,amount=1) 	#insert 1 empty row
	wb.active.cell(row=1, column=1).value = 'Endereço IP'
	wb.active.cell(row=1, column=2).value = 'Data e Hora'
	wb.active.cell(row=1, column=3).value = 'Resposta'

# Main Progran

# Create Sheet Names
wb.remove(wb['Sheet'])
length = 0
while length < list_length:  # Create sheet name
	get_file(length)
	sheet_name(length)
	length += 1

# Active Single Sheet
length = 0
while length < list_length:

	sheet_name = get_file(length)
	#print(sheet_name)
	act_sheet(sheet_name)
	ws.PAPERSIZE_A4
	column_dimensions()

# Paste Data
	file = open_file(length)
	IP = get_IP(file)
	date = get_date(file)
	status = get_status(file)
	for row in zip(IP, date, status):
		wb.active.append(row)
	#wb.active.insert_rows(1,amount=1) 	#insert 1 empty row 
	title()	# paste titles table
	length += 1

save()


# Se possivel converter a planilha para PDF
# Nome das colunas


